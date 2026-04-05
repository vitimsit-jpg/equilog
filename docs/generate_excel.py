import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Couleurs ──────────────────────────────────────────────────────────────────
C_HEADER_DARK   = "1C1C1E"
C_HEADER_MED    = "3A3A3C"
C_HEADER_LIGHT  = "F5F0E8"   # beige
C_GREEN         = "D4EDDA"
C_RED           = "F8D7DA"
C_ORANGE        = "FFE0B2"
C_YELLOW        = "FFF9C4"
C_GRAY          = "F0F0F0"
C_BLUE          = "D0E8FF"
C_WHITE         = "FFFFFF"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="888888")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, dark=True):
    cell.fill = fill(C_HEADER_DARK if dark else C_HEADER_MED)
    cell.font = font(bold=True, color="FFFFFF", size=10)
    cell.alignment = center()
    cell.border = border_thin()

def style_subheader(cell):
    cell.fill = fill(C_HEADER_LIGHT)
    cell.font = font(bold=True, color="1C1C1E", size=9)
    cell.alignment = center()
    cell.border = border_thin()

def style_cell(cell, bg=C_WHITE, bold=False, italic=False, color="000000"):
    cell.fill = fill(bg)
    cell.font = font(bold=bold, italic=italic, color=color, size=9)
    cell.alignment = center()
    cell.border = border_thin()

def style_label(cell, bg=C_GRAY):
    cell.fill = fill(bg)
    cell.font = font(bold=True, size=9)
    cell.alignment = left()
    cell.border = border_thin()


# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 1 — MATRICE PROFILS × FONCTIONNALITÉS
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1 – Profils × Features"

# Titre
ws1.merge_cells("A1:J1")
ws1["A1"] = "EQUILOG — Matrice Profils × Fonctionnalités"
ws1["A1"].fill = fill(C_HEADER_DARK)
ws1["A1"].font = font(bold=True, color="FFFFFF", size=12)
ws1["A1"].alignment = center()

# Légende ligne 2
ws1.merge_cells("A2:J2")
ws1["A2"] = "✅ Accès complet   🔶 Conditionnel   ❌ Bloqué   ⚠️ Non enforced (devrait être limité)"
ws1["A2"].fill = fill(C_HEADER_LIGHT)
ws1["A2"].font = font(italic=True, size=9)
ws1["A2"].alignment = left()

# En-têtes colonnes (profils)
headers = ["Fonctionnalité", "Catégorie", "loisir", "competition", "pro", "gerant", "+ module_coach", "+ module_gerant", "is_admin", "Notes"]
cols_bg  = [C_GRAY, C_GRAY, "D0F0C0", "C0D8FF", "FFD0A0", "E8C0FF", "FFE0E0", "C0F0F0", "FFD0D0", C_YELLOW]

for col, (h, bg) in enumerate(zip(headers, cols_bg), 1):
    c = ws1.cell(row=3, column=col, value=h)
    c.fill = fill(bg)
    c.font = font(bold=True, size=9)
    c.alignment = center()
    c.border = border_thin()

# Données : (fonctionnalité, catégorie, loisir, competition, pro, gerant, coach, gerant_mod, admin, notes)
rows = [
    # CORE
    ("Dashboard", "Core", "✅", "✅", "✅", "✅", "✅", "✅", "✅", ""),
    ("Carnet de santé", "Core", "✅", "✅", "✅", "✅", "✅", "✅", "✅", ""),
    ("Journal de travail", "Core", "✅", "✅", "✅", "✅", "✅", "✅", "✅", ""),
    ("Concours", "Core", "✅", "✅", "✅", "✅", "✅", "✅", "✅", ""),
    ("Budget", "Core", "✅", "✅", "✅", "✅", "✅", "✅", "✅", ""),
    ("Documents cheval", "Core", "✅", "✅", "✅", "✅", "✅", "✅", "✅", ""),
    ("Généalogie", "Core", "✅", "✅", "✅", "✅", "✅", "✅", "✅", ""),
    ("Historique vie", "Core", "✅", "✅", "✅", "✅", "✅", "✅", "✅", ""),
    # INDEX
    ("Horse Index (score)", "Horse Index", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "⚠️ Starter devrait être limité"),
    ("Classements", "Horse Index", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "⚠️ Devrait être Pro+"),
    ("Profil public /share", "Horse Index", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "Conditionnel : share_horse_index=true"),
    # PLANNING
    ("Planning hebdomadaire", "Planning", "✅", "✅", "✅", "✅", "✅", "✅", "✅", ""),
    ("Bibliothèque exercices", "Planning", "✅", "✅", "✅", "✅", "✅", "✅", "✅", ""),
    # IA
    ("AI Insights basiques", "IA", "✅", "✅", "✅", "✅", "✅", "✅", "✅", ""),
    ("Saisie vocale", "IA", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "Chrome/Safari uniquement"),
    ("CoachChat (streaming)", "IA", "❌", "❌", "❌", "❌", "✅", "❌", "✅", "Gate : module_coach=true"),
    ("Gestion élèves", "IA", "❌", "❌", "❌", "❌", "✅", "❌", "✅", "Gate : module_coach=true"),
    ("Checklist IA pré-concours", "IA", "✅", "✅", "✅", "✅", "✅", "✅", "✅", ""),
    # PREMIUM
    ("Video Analysis", "Premium", "⚠️❌", "⚠️❌", "⚠️❌", "⚠️❌", "⚠️❌", "⚠️❌", "✅", "Seule vraie restriction plan Starter"),
    ("Export PDF bilan annuel", "Premium", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "⚠️ Devrait être Pro+"),
    # NUTRITION
    ("Module Nutrition", "Modules", "🔶", "🔶", "🔶", "🔶", "🔶", "🔶", "✅", "Gate : horse.module_nutrition=true"),
    # GESTION
    ("Mon Écurie (/mon-ecurie)", "Gestion", "❌", "❌", "❌", "✅", "❌", "✅", "✅", "Gate : gerant OR module_gerant"),
    ("Communauté (feed)", "Social", "✅", "✅", "✅", "✅", "✅", "✅", "✅", ""),
    # ADMIN
    ("Back-office /admin", "Admin", "❌", "❌", "❌", "❌", "❌", "❌", "✅", "Gate : is_admin=true"),
]

COLOR_MAP = {
    "✅": C_GREEN,
    "❌": C_RED,
    "🔶": C_ORANGE,
    "⚠️❌": C_RED,
}
CURRENT_CAT = None

for r, row in enumerate(rows, 4):
    cat = row[1]
    if cat != CURRENT_CAT:
        CURRENT_CAT = cat
        # Séparateur de catégorie
        ws1.merge_cells(f"A{r}:J{r}")
        ws1[f"A{r}"] = f"  ▸ {cat}"
        ws1[f"A{r}"].fill = fill(C_HEADER_MED)
        ws1[f"A{r}"].font = font(bold=True, color="FFFFFF", size=9)
        ws1[f"A{r}"].alignment = left()
        r += 1
        # Re-ajouter la ligne de données après le séparateur
        # On ne peut pas faire ça facilement — simplement coller sans séparateur
        # → on va juste changer le bg de la col catégorie

    for col, val in enumerate(row, 1):
        c = ws1.cell(row=r, column=col, value=val)
        if col == 1:  # Fonctionnalité
            style_label(c, C_GRAY)
        elif col == 2:  # Catégorie
            c.fill = fill(C_HEADER_LIGHT)
            c.font = font(italic=True, size=8, color="555555")
            c.alignment = center()
            c.border = border_thin()
        elif col == 10:  # Notes
            c.fill = fill(C_YELLOW)
            c.font = font(italic=True, size=8)
            c.alignment = left()
            c.border = border_thin()
        else:
            bg = COLOR_MAP.get(str(val).strip(), C_WHITE)
            style_cell(c, bg=bg)

# Largeurs colonnes
ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 14
for col_letter in ["C", "D", "E", "F", "G", "H", "I"]:
    ws1.column_dimensions[col_letter].width = 14
ws1.column_dimensions["J"].width = 35
ws1.row_dimensions[1].height = 22
ws1.row_dimensions[2].height = 16
ws1.freeze_panes = "C4"


# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 2 — MATRICE ABONNEMENTS × FONCTIONNALITÉS
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2 – Abonnements × Features")

ws2.merge_cells("A1:G1")
ws2["A1"] = "EQUILOG — Matrice Abonnements × Fonctionnalités (état réel du code)"
ws2["A1"].fill = fill(C_HEADER_DARK)
ws2["A1"].font = font(bold=True, color="FFFFFF", size=12)
ws2["A1"].alignment = center()

ws2.merge_cells("A2:G2")
ws2["A2"] = "⚠️ ATTENTION : le paywall est quasi inexistant dans le code. Seule Video Analysis est réellement bloquée en Starter."
ws2["A2"].fill = fill("FFE0B2")
ws2["A2"].font = font(bold=True, size=9, color="7F4000")
ws2["A2"].alignment = left()

# Headers
plan_headers = ["Fonctionnalité", "Starter\n(gratuit)", "Pro\n(9,90€/mois)", "Écurie\n(29€/mois)", "Enforced ?", "Devrait être", "Notes"]
plan_bg      = [C_GRAY, "E8F5E9", "E3F2FD", "F3E5F5", C_RED, C_ORANGE, C_YELLOW]

for col, (h, bg) in enumerate(zip(plan_headers, plan_bg), 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.fill = fill(bg)
    c.font = font(bold=True, size=9)
    c.alignment = center()
    c.border = border_thin()

plan_rows = [
    ("Dashboard + Core modules",       "✅", "✅", "✅", "✅ Oui",  "Tous",        ""),
    ("Carnet santé / Training / Budget","✅", "✅", "✅", "✅ Oui",  "Tous",        ""),
    ("Horse Index (score)",             "✅", "✅", "✅", "❌ Non",  "Pro+",        "⚠️ Starter devrait avoir Horse Index limité"),
    ("Classements",                     "✅", "✅", "✅", "❌ Non",  "Pro+",        ""),
    ("AI Insights",                     "✅", "✅", "✅", "❌ Non",  "Pro+",        ""),
    ("Export PDF",                      "✅", "✅", "✅", "❌ Non",  "Pro+",        ""),
    ("Profil public /share",            "✅", "✅", "✅", "❌ Non",  "Pro+",        ""),
    ("CoachChat",                       "✅", "✅", "✅", "❌ Non",  "Pro+",        "Gated par module_coach, pas par plan"),
    ("Saisie vocale",                   "✅", "✅", "✅", "❌ Non",  "Pro+",        "Chrome/Safari"),
    ("VIDEO ANALYSIS",                  "❌", "✅", "✅", "✅ OUI",  "Pro+",        "⭐ SEULE vraie restriction active"),
    ("Dashboard gérant /mon-ecurie",    "✅", "✅", "✅", "❌ Non",  "Écurie",      "Gated par profil/module, pas par plan"),
    ("Multi-cavaliers",                 "✅", "✅", "✅", "❌ Non",  "Écurie",      ""),
    ("Chevaux illimités",               "✅", "✅", "✅", "❌ Non",  "Starter=1",   "PLAN_LIMITS = Infinity pour tous"),
    ("Module Nutrition",                "✅", "✅", "✅", "❌ Non",  "—",           "Toggle par cheval"),
    ("Communauté",                      "✅", "✅", "✅", "❌ Non",  "Tous",        ""),
]

for r, row in enumerate(plan_rows, 4):
    feat, s, p, e, enf, should, note = row
    row_data = [feat, s, p, e, enf, should, note]
    for col, val in enumerate(row_data, 1):
        c = ws2.cell(row=r, column=col, value=val)
        if col == 1:
            style_label(c, C_GRAY)
        elif col in [2, 3, 4]:
            bg = C_GREEN if val == "✅" else (C_RED if val == "❌" else C_ORANGE)
            style_cell(c, bg=bg)
        elif col == 5:
            bg = C_GREEN if "Oui" in str(val) else C_RED
            style_cell(c, bg=bg, bold=("OUI" in str(val)))
        elif col == 6:
            style_cell(c, bg=C_ORANGE)
        else:
            c.fill = fill(C_YELLOW)
            c.font = font(italic=True, size=8)
            c.alignment = left()
            c.border = border_thin()

ws2.column_dimensions["A"].width = 30
for l in ["B","C","D"]:
    ws2.column_dimensions[l].width = 14
ws2.column_dimensions["E"].width = 12
ws2.column_dimensions["F"].width = 14
ws2.column_dimensions["G"].width = 38
ws2.freeze_panes = "B4"


# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 3 — PARCOURS PRINCIPAUX PAR PROFIL
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3 – Parcours par Profil")

ws3.merge_cells("A1:H1")
ws3["A1"] = "EQUILOG — Parcours Principaux par Profil"
ws3["A1"].fill = fill(C_HEADER_DARK)
ws3["A1"].font = font(bold=True, color="FFFFFF", size=12)
ws3["A1"].alignment = center()

flow_headers = ["Profil", "Objectif principal", "Pages prioritaires", "Actions principales",
                "Module requis", "Dépendance plan", "Parcours secondaires", "Risque / Limite"]
for col, h in enumerate(flow_headers, 1):
    c = ws3.cell(row=2, column=col, value=h)
    style_header(c)
    c.font = font(bold=True, color="FFFFFF", size=9)

flow_rows = [
    ("🟢 LOISIR",
     "Suivi quotidien, tranquillité d'esprit",
     "/dashboard\n/health\n/training\n/budget",
     "Logger soins\nLogger séances\nSuivre dépenses\nVoir score HI",
     "Aucun obligatoire\n(module_coach optionnel)",
     "Video bloquée si Starter\n(seule vraie limite)",
     "Généalogie\nCommunauté\nDocuments",
     "Peu d'urgence à passer Pro"),

    ("🔵 COMPETITION",
     "Performance, gestion concours",
     "/competitions\n/training\n/planning\n/classements",
     "Créer concours\nChecklist IA\nPlanifier séances\nSuivre classement",
     "Aucun obligatoire",
     "Video (Pro) utile pour analyse\nIndex accessible en Starter",
     "Budget\nSanté post-effort\n/share (palmarès)",
     "Pas de vrai mur avant Video"),

    ("🟠 PRO",
     "Performance intensive\nMulti-chevaux",
     "/training\n/video\n/dashboard\n/horses/[id]",
     "Logs détaillés\nSaisie vocale\nAnalyse vidéo\nSuivi performance",
     "Aucun obligatoire\n(module_coach probable)",
     "Video = feature centrale\n→ Seul profil vraiment bloqué par Starter",
     "Concours\nSanté",
     "Conversion Pro très liée\nà Video Analysis"),

    ("🟣 GERANT",
     "Gestion structure écurie\nSupervision cheptel",
     "/mon-ecurie\n/health\n/budget\n/dashboard",
     "Vue multi-chevaux\nSuivi soins cheptel\nGestion budget\nRappels automatiques",
     "module_gerant=true\nOU profile_type=gerant",
     "/mon-ecurie accessible\nsans plan Écurie\n(non enforced)",
     "Planning\nClassements écurie",
     "Friction si onboarding\nne configure pas le module"),

    ("🤖 COACH\n(module_coach)",
     "Suivi élèves\nRecommandations IA",
     "CoachChat (toutes pages /horses)\nTraining\nHorse Index",
     "Chat contextuel\nSuivi progression\nAI Insights avancés",
     "module_coach=true\n(non lié à un plan)",
     "Aucune restriction plan\npour ce module",
     "Gestion élèves\nExport PDF",
     "Module non monétisé\n→ accessible gratuitement"),

    ("⚙️ ADMIN\n(is_admin)",
     "Monitoring plateforme\nSupport",
     "/admin\n/admin/users\n/admin/audit",
     "Voir analytics\nGérer users\nSuspendre comptes",
     "is_admin=true",
     "Bypasse tout",
     "Toutes les pages",
     "Usage interne uniquement"),
]

profile_colors = ["D4EDDA", "D0E8FF", "FFE0B2", "F3D0FF", "FFD0D0", "E8E8E8"]

for r, (row, bg) in enumerate(zip(flow_rows, profile_colors), 3):
    for col, val in enumerate(row, 1):
        c = ws3.cell(row=r, column=col, value=val)
        if col == 1:
            c.fill = fill(bg)
            c.font = font(bold=True, size=10)
            c.alignment = center()
            c.border = border_thin()
        elif col in [6, 8]:
            c.fill = fill(C_ORANGE if col == 6 else C_RED)
            c.font = font(size=8, italic=(col==8))
            c.alignment = left()
            c.border = border_thin()
        else:
            c.fill = fill(C_WHITE)
            c.font = font(size=9)
            c.alignment = left()
            c.border = border_thin()

ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 22
ws3.column_dimensions["C"].width = 20
ws3.column_dimensions["D"].width = 22
ws3.column_dimensions["E"].width = 20
ws3.column_dimensions["F"].width = 22
ws3.column_dimensions["G"].width = 20
ws3.column_dimensions["H"].width = 25
for r in range(3, 9):
    ws3.row_dimensions[r].height = 60
ws3.freeze_panes = "B3"


# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 4 — RISQUES ET ZONES FLOUES
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4 – Risques & Zones Floues")

ws4.merge_cells("A1:F1")
ws4["A1"] = "EQUILOG — Risques Produit & Zones à Clarifier"
ws4["A1"].fill = fill(C_HEADER_DARK)
ws4["A1"].font = font(bold=True, color="FFFFFF", size=12)
ws4["A1"].alignment = center()

risk_headers = ["#", "Risque / Problème", "Type", "Sévérité", "Détail", "Action recommandée"]
risk_bg = [C_GRAY, C_GRAY, C_GRAY, C_GRAY, C_GRAY, C_YELLOW]
for col, (h, bg) in enumerate(zip(risk_headers, risk_bg), 1):
    c = ws4.cell(row=2, column=col, value=h)
    c.fill = fill(C_HEADER_DARK)
    c.font = font(bold=True, color="FFFFFF", size=9)
    c.alignment = center()
    c.border = border_thin()

SEVERITY = {
    "🔴 Critique":  "FF4C4C",
    "🟡 Important": "FFC107",
    "🟢 Faible":    "4CAF50",
}

risk_rows = [
    ("R1", "Paywall non enforced\n→ zéro conversion payante",
     "Business", "🔴 Critique",
     "Tous users ont plan 'ecurie' par défaut (migration 054).\nPLAN_LIMITS = Infinity pour tous.",
     "Décider si/quand activer les restrictions.\nPrévoir migration users existants."),

    ("R2", "Page /pricing promet des features non livrées",
     "Trust", "🔴 Critique",
     "Starter annoncé sans Horse Index, sans AI, 1 cheval.\nRéalité : tout accessible.",
     "Aligner pricing page avec réalité\nOU implémenter les vrais gates."),

    ("R3", "Video Analysis = seule vraie raison de passer Pro\n→ très fragile",
     "Business", "🔴 Critique",
     "Si un user n'a pas besoin de video,\nil n'a aucun incentive à payer.",
     "Ajouter d'autres gates de valeur.\nOU repositionner Video comme feature phare."),

    ("R4", "Double système de profils\n(ProfileType + UserType legacy)",
     "Technique", "🟡 Important",
     "Sidebar utilise les 6 UserType legacy.\nReste du code utilise 4 ProfileType.\nMaintenance double.",
     "Migrer complètement vers ProfileType.\nSupprimer les checks legacy."),

    ("R5", "module_coach non lié à un plan\n→ non monétisé",
     "Business", "🟡 Important",
     "module_coach=true donne accès au CoachChat\nsans nécessiter un abonnement Pro.",
     "Lier module_coach au plan Pro dans le gate.\nOU en faire une option payante séparée."),

    ("R6", "horse_user_roles sans UI de gestion",
     "Technique", "🟡 Important",
     "Table en DB, hooks implémentés,\nmais aucun écran pour assigner un rôle.",
     "Implémenter l'UI OU supprimer si non prioritaire."),

    ("R7", "Données rider profile collectées, non exploitées",
     "UX", "🟡 Important",
     "7 champs rider_* (asymétrie, pathologies, zones douloureuses...)\ncollectés à l'onboarding sans usage visible.",
     "Soit exploiter dans les AI recommandations,\nsoit supprimer des étapes onboarding."),

    ("R8", "7 étapes onboarding\n→ drop rate potentiel",
     "Activation", "🟡 Important",
     "Steps 5 (trousseau) et 6 (rider profile)\ntrès détaillées avant toute valeur perçue.",
     "A/B test onboarding court (3 steps) vs long.\nDifférer les steps optionnelles."),

    ("R9", "Navigation réordonnée ≠ restriction d'accès",
     "UX", "🟢 Faible",
     "Profils personnalisent l'ordre sidebar\nmais tout est accessible par URL directe.",
     "Acceptable si pas de vraie restriction prévue.\nDocumenter l'intention."),

    ("R10", "Mon Écurie lié au profil, pas au plan",
     "Business", "🟡 Important",
     "Accès /mon-ecurie via profil=gerant OU module_gerant.\nPas via plan=ecurie.",
     "Décider si Mon Écurie = feature plan Écurie\net aligner le gate en conséquence."),
]

for r, row in enumerate(risk_rows, 3):
    num, prob, typ, sev, detail, action = row
    row_data = [num, prob, typ, sev, detail, action]
    sev_color = SEVERITY.get(sev, "FFFFFF")
    for col, val in enumerate(row_data, 1):
        c = ws4.cell(row=r, column=col, value=val)
        if col == 4:
            c.fill = fill(sev_color)
            c.font = font(bold=True, size=9, color="FFFFFF")
            c.alignment = center()
        elif col == 1:
            c.fill = fill(C_GRAY)
            c.font = font(bold=True, size=9)
            c.alignment = center()
        elif col == 5:
            c.fill = fill(C_WHITE)
            c.font = font(size=8, italic=True)
            c.alignment = left()
        elif col == 6:
            c.fill = fill(C_YELLOW)
            c.font = font(size=8)
            c.alignment = left()
        else:
            c.fill = fill(C_WHITE)
            c.font = font(size=9, bold=(col==2))
            c.alignment = left()
        c.border = border_thin()

ws4.column_dimensions["A"].width = 5
ws4.column_dimensions["B"].width = 32
ws4.column_dimensions["C"].width = 12
ws4.column_dimensions["D"].width = 14
ws4.column_dimensions["E"].width = 38
ws4.column_dimensions["F"].width = 35
for r in range(3, 14):
    ws4.row_dimensions[r].height = 45
ws4.freeze_panes = "B3"


# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 5 — QUESTIONS À CLARIFIER
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5 – Questions à Clarifier")

ws5.merge_cells("A1:E1")
ws5["A1"] = "EQUILOG — Décisions à prendre avant de développer"
ws5["A1"].fill = fill(C_HEADER_DARK)
ws5["A1"].font = font(bold=True, color="FFFFFF", size=12)
ws5["A1"].alignment = center()

q_headers = ["#", "Question", "Contexte", "Options possibles", "Urgence"]
for col, h in enumerate(q_headers, 1):
    c = ws5.cell(row=2, column=col, value=h)
    style_header(c)

q_rows = [
    ("Q1", "Quel est le vrai modèle paywall ?",
     "La page /pricing définit des restrictions\nqui ne sont pas implémentées dans le code.",
     "A) Implémenter les vraies restrictions\nB) Simplifier la page pricing\nC) Freemium sans restrictions réelles",
     "🔴 Bloquant"),

    ("Q2", "6 profils ou 4 profils ?",
     "UserType (6 valeurs legacy) et ProfileType (4 valeurs)\ncoexistent dans le code.",
     "A) Migrer complètement vers 4 ProfileType\nB) Garder les 6 comme distinction UI\nC) Ajouter profils manquants au nouveau système",
     "🟡 Important"),

    ("Q3", "module_coach = feature gratuite ou payante ?",
     "Actuellement activable sans abonnement Pro.\nCoachChat est une feature IA premium.",
     "A) Gratuit (outil de rétention)\nB) Lié au plan Pro (monétisation)\nC) Module additionnel payant séparé",
     "🟡 Important"),

    ("Q4", "Mon Écurie = plan Écurie ou profil gerant ?",
     "/mon-ecurie est accessible via profil OU module,\npas via le plan. Incohérence avec /pricing.",
     "A) Garder lié au profil (logique métier)\nB) Lier au plan Écurie (logique business)\nC) Les deux (profil OU plan)",
     "🟡 Important"),

    ("Q5", "Que faire de horse_user_roles ?",
     "Table en DB, hooks présents, zéro UI.\nFeature zombie.",
     "A) Implémenter l'UI de gestion des rôles\nB) Supprimer (trop tôt)\nC) Documenter pour v3",
     "🟢 Faible"),

    ("Q6", "Les données rider_* ont-quel usage ?",
     "Collectées en onboarding (niveau, asymétrie,\npathologies...) mais non visibles dans l'app.",
     "A) Personnaliser les recommandations IA\nB) Afficher dans le profil cavalier\nC) Supprimer des étapes onboarding",
     "🟡 Important"),

    ("Q7", "Onboarding : 7 steps ou raccourcir ?",
     "7 étapes avec des steps très détaillées\navant que l'user voie de la valeur.",
     "A) Garder 7 steps (données complètes)\nB) Onboarding court + complétion progressive\nC) Steps optionnelles différées",
     "🟡 Important"),

    ("Q8", "ICr milestones : dynamiques ou manuels ?",
     "Les milestones de croissance poulain existent\nmais leur calcul par rapport à birth_year n'est pas clair.",
     "A) Calculer automatiquement depuis birth_year\nB) Saisie manuelle par l'utilisateur\nC) Proposer des dates suggérées (IA)",
     "🟢 Faible"),
]

URGENCE_COLORS = {
    "🔴 Bloquant":  "FF4C4C",
    "🟡 Important": "FFC107",
    "🟢 Faible":    "4CAF50",
}

for r, row in enumerate(q_rows, 3):
    num, q, ctx, opts, urg = row
    for col, val in enumerate([num, q, ctx, opts, urg], 1):
        c = ws5.cell(row=r, column=col, value=val)
        if col == 1:
            c.fill = fill(C_GRAY)
            c.font = font(bold=True, size=9)
            c.alignment = center()
        elif col == 5:
            urg_col = URGENCE_COLORS.get(val, "FFFFFF")
            c.fill = fill(urg_col)
            c.font = font(bold=True, size=9, color="FFFFFF")
            c.alignment = center()
        elif col == 4:
            c.fill = fill(C_BLUE)
            c.font = font(size=8)
            c.alignment = left()
        elif col == 3:
            c.fill = fill(C_WHITE)
            c.font = font(size=8, italic=True)
            c.alignment = left()
        else:
            c.fill = fill(C_WHITE)
            c.font = font(size=9, bold=(col==2))
            c.alignment = left()
        c.border = border_thin()

ws5.column_dimensions["A"].width = 5
ws5.column_dimensions["B"].width = 30
ws5.column_dimensions["C"].width = 35
ws5.column_dimensions["D"].width = 38
ws5.column_dimensions["E"].width = 14
for r in range(3, 12):
    ws5.row_dimensions[r].height = 55
ws5.freeze_panes = "B3"


# ══════════════════════════════════════════════════════════════════════════════
# SAUVEGARDE
# ══════════════════════════════════════════════════════════════════════════════
output_path = "/Users/vincenttimsit/equilog/docs/equilog-product-cartography.xlsx"
wb.save(output_path)
print(f"✅ Fichier créé : {output_path}")
